import pandas as pd
import numpy as np
import math
import matplotlib.pyplot as plt

from copy import deepcopy
from openpyxl import load_workbook  # conda install openpyxl

turbine_data={
    'sub_system': 'Wind energy source & Transport',
    'element': 'Offshore wind park',
    'component': 'Turbine',
    'escalation_base_year': 2030,
    'escalation_rate': 0.02,
    'capex_per_unit': 1_495_000,
    'capex_per_unit_units': 'EUR/MW',
    'unit': 3_000,
    'unit_units': 'MW',
    'construction_duration': 3,
    'share_of_investments': [0.4, 0.3, 0.3],
    'economic_lifetime': 25,
    'depreciation_flag': 1,
    'depreciation_rate': 0.01,
    'yearly_variable_costs_flag': 1,
    'yearly_variable_costs_rate': 0.03,
    'insurance_flag': 1,
    'insurance_rate': 0.005,
    'decommissioning_rate': 0.02,
    'residual_value': 0.01}

foundation_data={
    'sub_system': 'Wind energy source & Transport',
    'element': 'Offshore wind park',
    'component': 'Foundation & cable',
    'escalation_base_year': 2030,
    'escalation_rate': 0.02,
    'capex_per_unit': 2_691_000,
    'capex_per_unit_units': 'EUR/MW',
    'unit': 3_000,
    'unit_units': 'MW',
    'construction_duration': 3,
    'share_of_investments': [0.4, 0.3, 0.3],
    'economic_lifetime': 50,
    'depreciation_flag': 1,
    'depreciation_rate': 0.01,
    'yearly_variable_costs_flag': 1,
    'yearly_variable_costs_rate': 0.03,
    'insurance_flag': 1,
    'insurance_rate': 0.005,
    'decommissioning_rate': 0.02,
    'residual_value': 0.01}


class CashflowProperties(object):

    def __init__(self,
                 sub_system,
                 element,
                 component,
                 escalation_base_year,
                 escalation_rate,
                 capex_per_unit,
                 capex_per_unit_units,
                 unit,
                 unit_units,
                 construction_duration,
                 share_of_investments,
                 economic_lifetime,
                 depreciation_flag,
                 depreciation_rate,
                 yearly_variable_costs_flag,
                 yearly_variable_costs_rate,
                 insurance_flag,
                 insurance_rate,
                 decommissioning_rate,
                 residual_value,
                 *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.sub_system = sub_system
        self.element = element
        self.component = component
        self.escalation_base_year = escalation_base_year
        self.escalation_rate = escalation_rate
        self.capex_per_unit = capex_per_unit
        self.capex_per_unit_units = capex_per_unit_units
        self.unit = unit
        self.unit_units = unit_units
        self.construction_duration = construction_duration
        self.share_of_investments = share_of_investments
        self.economic_lifetime = economic_lifetime
        self.depreciation_flag = depreciation_flag
        self.depreciation_rate = depreciation_rate
        self.yearly_variable_costs_flag = yearly_variable_costs_flag
        self.yearly_variable_costs_rate = yearly_variable_costs_rate
        self.insurance_flag = insurance_flag
        self.insurance_rate = insurance_rate
        self.decommissioning_rate = decommissioning_rate
        self.residual_value = residual_value

        self.investment_years = []
        self.summed_escalated_capex = []
        self.divestment_years = []
        self.divestment_values = []
        self.decommissioning_years = []
        self.decommissioning_values = []

    def generate_cashflows(self,
                           startyear: int = 2030,
                           lifecycle: int = 11,
                           debug: bool = False):
        """
        verschillende Opex periodes telkens op basis de summed escalated capex van de laatste (re)investment
        Opex van de vorige periode loopt door gedurende de herinvesterings capex
        divestment als positieve Capex (niet escalated, waarde gebaseerd op de summed escalated capex van de laatste (re)investment)
        decommissioning als Opex (escalated, waarde gebaseerd op de summed escalated capex van de laatste (re)investment)

        startyear: the year when the first CAPEX investment will be scheduled
        lifecycle: the overall lifecycle of the project
        debug: True: show messages, False: hide messages
        """

        def escalate_list(list_years, list_values):
            for i, list_year in enumerate(list_years):
                list_values[i] = list_values[i] * self.escalation_values[
                    [index for index, escalation_year in enumerate(self.escalation_years) if
                     escalation_year == list_year][0]]
            return list_values

        # initialise  lists
        self.escalation_years = []
        self.escalation_values = []

        capex_years = []
        capex_values = []
        opex_years = []
        opex_values = []
        revenue_years = []
        revenue_values = []

        # --------------------------------------------------------------------------------------------------------------

        # 1. generate a list of escalation factors
        previous = 1
        for index, year in enumerate(list(range(self.escalation_base_year, self.escalation_base_year + lifecycle))):
            # should be "self.escalation_base_year + lifecycle - 1" but range omits the last value (so -1 can be left out)
            previous = previous * (1 + self.escalation_rate)
            self.escalation_years.append(year)
            self.escalation_values.append(previous)

        # --------------------------------------------------------------------------------------------------------------

        if debug:
            print('')
            print('*** determine investment_years, capex_years and capex_values ****')
            print('')

        # 2. loop through years from startyear to escalation_base_year + lifecycle: find (re)invest_years and fill in capex
        year = startyear
        insufficient_reinvestment_time = False
        while year <= self.escalation_base_year + lifecycle - 1:

            if not self.investment_years or year == self.investment_years[-1] + self.economic_lifetime:
                # the first action will always be to invest because per definition startyear is the first investment_year
                # then invest each time year == investment_years[-1] + economic_lifetime

                # add investment years (per definition startyear is the first investment_year)
                if not self.investment_years:
                    # add the first investment year ...
                    self.investment_years.append(year)

                    # ... add the (re)investment capex to the capex_years and capex_values
                    for i in range(self.construction_duration):
                        capex_years.append(self.investment_years[-1] + i)
                        capex_values.append(-self.capex_per_unit * self.unit * self.share_of_investments[i])

                    if debug:
                        print('initial investment in {}'.format(year))
                        print('')

                else:
                    if year + self.construction_duration < self.escalation_base_year + lifecycle:
                        # if there is sufficient time to implement the (re)investment ...
                        self.investment_years.append(year)

                        # ... add the (re)investment capex to the capex_years and capex_values
                        for i in range(self.construction_duration):
                            capex_years.append(self.investment_years[-1] + i)
                            capex_values.append(-self.capex_per_unit * self.unit * self.share_of_investments[i])

                        if debug:
                            print('reinvestment in {}'.format(year))
                            print('')
                    else:
                        # ... else trigger the insufficient_reinvestment_time flag and do not add any Capex values
                        insufficient_reinvestment_time = True
                        if debug:
                            print('not enough time to reinvest')
                            print('')

            # increase the year by one
            year = year + 1

        # --------------------------------------------------------------------------------------------------------------

        # 3. determine the summed_escalated_capex per investment round
        for index, investment_year in enumerate(self.investment_years):
            # find the Capex values associated with the last investment (for summed_escalated_capex)
            bools = [item >= self.investment_years[index] and
                     item <= self.investment_years[index] + self.construction_duration for
                     item in capex_years]
            temp_capex_years = deepcopy([a for a, b in zip(capex_years, bools) if b])
            temp_capex_values = deepcopy([a for a, b in zip(capex_values, bools) if b])

            # proceed to escalate these
            temp_capex_values = escalate_list(temp_capex_years, temp_capex_values)

            # store the summed_escalated_capex value
            self.summed_escalated_capex.append(sum(temp_capex_values))

        # --------------------------------------------------------------------------------------------------------------

        # 4. escalate the CAPEX using the list of escalation factors
        capex_values = escalate_list(capex_years, capex_values)

        # --------------------------------------------------------------------------------------------------------------

        # 5. add divestment (no escalation)
        for index, investment_year in enumerate(self.investment_years):
            # divestment_year
            divestment_year = min([investment_year + self.construction_duration + self.economic_lifetime - 1,
                                   self.escalation_base_year + lifecycle - 1])

            # divestment_value
            # if the economic life ends before the total project lifecycle ...
            if investment_year + self.construction_duration + self.economic_lifetime < self.escalation_base_year + lifecycle:
                # diminish the summed_escalated_capex with number of years times to lifetime end, times the depreciation (should typically be zero)
                divestment_value = -1 * (self.summed_escalated_capex[index] -
                                         self.summed_escalated_capex[index] * self.depreciation_rate * (
                                             (self.investment_years[index] + self.construction_duration + self.economic_lifetime - 1) -
                                             (self.investment_years[index] + self.construction_duration - 1)))
            else:
                # diminish the summed_escalated_capex with number of years to lifecycle end, times the depreciation
                divestment_value = -1 * (self.summed_escalated_capex[index] -
                                         self.summed_escalated_capex[index] * self.depreciation_rate * (
                                             (self.escalation_base_year + lifecycle - 1) -
                                             (self.investment_years[index] + self.construction_duration - 1)))

            # add to capex_values ...
            if divestment_year in capex_years:
                # ... if divestment_year is already in capex_years add the value to the existing value ...
                value_index = [index for index, capex_year in enumerate(capex_years) if divestment_year == capex_year][0]
                capex_values[value_index] = capex_values[value_index] + divestment_value
            else:
                # ... otherwise append the year and value as new entries to the capex lists
                capex_years.append(divestment_year)
                capex_values.append(divestment_value)

            # save info to object, for later inspection
            self.divestment_years.append(divestment_year)
            self.divestment_values.append(divestment_value)

            if debug:
                print('divestment year: {}'.format(divestment_year))
                print('divestment value: {:.2f}'.format(divestment_value))
                print('')

        if debug:
            # note that the divestment_value should not be escalated
            print('CAPEX years escalated: {}'.format(capex_years))
            print('CAPEX values escalated: {}'.format(capex_values))

        # --------------------------------------------------------------------------------------------------------------

        if debug:
            print('')
            print('*** determine opex_years and opex_values (including decommissioning) ****')
            print('')

        # 6. add Opex values
        # for each investment_year ...
        for index, investment_year in enumerate(self.investment_years):
            # ... determine opex_value
            opex_value = self.summed_escalated_capex[index] * (self.yearly_variable_costs_rate + self.insurance_rate)

            if debug:
                print('investment range start {} stop {}'.format(
                    investment_year + self.construction_duration,
                    min([investment_year + self.construction_duration + self.economic_lifetime,
                         self.escalation_base_year + lifecycle]) - 1))
                print('opex_value {:.2f}, based on a summed_escalated_capex of {:.2f} '.format(
                    opex_value,
                    self.summed_escalated_capex[index]))
                print('decommissioning cost of {:.2f} added in {}, based on a summed_escalated_capex of {:.2f} '.format(
                    self.summed_escalated_capex[index] * self.decommissioning_rate,
                    min([investment_year + self.construction_duration + self.economic_lifetime - 1,
                         self.escalation_base_year + lifecycle - 1]), self.summed_escalated_capex[index]))
                print('')

            # ... and, determine the investment cycle time interval and add the opex_value to that interval
            opex_years_inv_cycle = list(range(
                investment_year + self.construction_duration,
                min([investment_year + self.construction_duration + self.economic_lifetime,
                     self.escalation_base_year + lifecycle])))
            opex_values_inv_cycle = [opex_value] * len(opex_years_inv_cycle)

            # ... and, add the decommissioning value to the last field of the investment cycle time interval
            opex_values_inv_cycle[-1] = opex_values_inv_cycle[-1] + self.summed_escalated_capex[
                index] * self.decommissioning_rate

            # save info to object, for later inspection
            self.decommissioning_years.append(min([investment_year + self.construction_duration + self.economic_lifetime - 1,
                         self.escalation_base_year + lifecycle - 1]))
            self.decommissioning_values.append(self.summed_escalated_capex[index] * self.decommissioning_rate)

            # ... add these opex_values for the investment cycle time interval to the overall opex list (years and values)
            opex_years = opex_years + opex_years_inv_cycle
            opex_values = opex_values + opex_values_inv_cycle

        # 7. escalate the OPEX using the list of escalation factors
        opex_values = escalate_list(opex_years, opex_values)

        if debug:
            print('OPEX years escalated: {}'.format(opex_years))
            print('OPEX values escalated: {}'.format(opex_values))

        # --------------------------------------------------------------------------------------------------------------

        # 8. escalate the revenues using the list of escalation factors
        # revenue_years and revenue_values are for now still empty lists
        for i, revenue_year in enumerate(revenue_years):
            revenue_values[i] = revenue_values[i] * escalation_list[
                [index for index, escalation_year in enumerate(escalation_years) if escalation_year == revenue_year][0]]

        # --------------------------------------------------------------------------------------------------------------

        # 9. create dataframe with cashflows
        # use the _years and _values lists completed above to create a dataframe and add it to the object for later use
        self.df = create_cashflow_dataframe(escalation_base_year=self.escalation_base_year,
                                            lifecycle=lifecycle,
                                            capex={'years': capex_years,
                                                   'values': capex_values},
                                            opex={'years': opex_years,
                                                  'values': opex_values},
                                            revenue={'years': revenue_years,
                                                     'values': revenue_values})


Object = type('Object', (
        CashflowProperties,
    ), {})


def test_dataframe(df):
    """This method performs a set of tests on a dataframe to see if it has the right properties

    Note that it check not only the presence of columns but also their spelling

    Todo: add other tests if needed
    """

    # assert dataframe contain the required fields
    assert df.index.name == 'years', f"expected index.name 'years' not found"
    assert 'years' in df.columns, f"expected column 'years' not found"
    assert 'capex' in df.columns, f"expected column 'capex' not found"
    assert 'opex' in df.columns, f"expected column 'opex' not found"
    assert 'revenue' in df.columns, f"expected column 'revenue' not found"

def create_cashflow_dataframe(escalation_base_year=2030, lifecycle=50,
                              capex={'years': [2001, 2002],
                                     'values': [-5_000_000, -5_000_000]},
                              opex={'years': list(range(2003, 2011)),
                                    'values': 8 * [-300_000]},
                              revenue={'years': list(range(2003, 2011)),
                                       'values': 8 * [1_500_000]}):
    """This method returns a dataframe with 'years' as index and index_name and columns years, capex, opex and revenue.

    The method first spans up a list of years based on 'startyear' and 'lifecycle'
    Then it initialises the other required columns: 'capex', 'opex' and 'revenue'
    Next it places the years-values combinations of 'capex', 'opex' and 'revenue' on the right places in de dataframe

    The end result is a dataframe that shows the capital expenses for each year

    """

    df = pd.DataFrame()

    # create list of years using startyear and lifecycle and set years as index
    years = list(range(escalation_base_year, escalation_base_year + lifecycle))
    df['years'] = years
    df.index = years
    df.index.name = 'years'

    # initialise capex, opex and revenue as zero
    df['capex'] = 0
    df['opex'] = 0
    df['revenue'] = 0

    # add capex from input
    for index, year in enumerate(capex['years']):
        df.loc[year, 'capex'] = capex['values'][index]

    # add opex from input
    for index, year in enumerate(opex['years']):
        df.loc[year, 'opex'] = opex['values'][index]

    # add revenue from input
    for index, year in enumerate(revenue['years']):
        df.loc[year, 'revenue'] = revenue['values'][index]

    # assert that dataframe adheres to prescribed standards
    test_dataframe(df)

    return df

def combine_cashflow_dataframes(dfs):
    """We assume that dfs is a list of dataframes that has a capex, opex and revenue column and years as index

    We add all years to a list 'years'. Next we determine the min and max year in that combined list.
    We create a new dataframe named df_combined that has all available years as a column and set as index.
    Next we step through the list of dataframes and add one by one all values to df_combined.

    Finally we return the combined dataframe.

    Todo: see if it is useful to also add a base year, so that you can calculate 
    s to a give baseyear
    """

    # assert all dataframes contain the required fields
    for df in dfs:
        test_dataframe(df)

    years = []
    for df in dfs:
        years = years + df.years.tolist()

    min_year = min(years)
    max_year = max(years)

    new_years = list(range(min_year, max_year + 1))

    df_combined = pd.DataFrame()
    df_combined['years'] = new_years
    df_combined['capex'] = 0
    df_combined['opex'] = 0
    df_combined['revenue'] = 0
    df_combined.index = new_years
    df_combined.index.name = 'years'

    for df in dfs:
        for year in df.years.tolist():
            df_combined.loc[year, 'capex'] = (df_combined['capex'].loc[year] + df['capex'].loc[year]).copy()
            df_combined.loc[year, 'opex'] = (df_combined['opex'].loc[year] + df['opex'].loc[year]).copy()
            df_combined.loc[year,'revenue'] = (df_combined['revenue'].loc[year] + df['revenue'].loc[year]).copy()

    return df_combined


def calculate_npv(df, escalation_base_year, WACC=0.07):
    """This method expects a dataframe that has years as index and index_name, and at least has columns
    named years, capex, opex, revenue.

    The method sums up all cashflows per year and adds these as a separate columns
    Also a cumulative cashflow column is added
    Next the npv is calculated
    Also a cumlative npv column is added

    Todo: see if it is useful to also add a 'baseyear', so that you can calculate npvs to a given 'baseyear'
    """

    # assert that dataframe adheres to prescribed standards
    test_dataframe(df)

    # collect the cashflows and add a 'cashflow' column
    df['cashflow'] = df.capex.copy() + df.opex.copy() + df.revenue.copy()

    # add the cumsum of cashflows to the 'cashflow_sum' column
    df['cashflow_sum'] = df['cashflow'].cumsum()

    # intitialise the 'npv' column with zeros
    df['npv'] = 0

    # calculate the npv through the years from the 2nd year up to the end and add the values to the 'npv' column
    for year in df.years.tolist()[:]:
        df.loc[year, 'npv'] = df['cashflow'].loc[year] * (1 / ((1 + WACC) ** (year - escalation_base_year)))

    # add the cumsum of npvs to the 'npv_sum' column
    df['npv_sum'] = df['npv'].cumsum()

    return df


def create_npv_plot(df, title=r'CAPEX, OPEX and Revenues and NPV', fname=r'test.png', x1=0, y1=0, x2=0, y2=0,
                    cash_flow_lims=[], npv_lims=[]):
    """This method creates a basic plot"""

    def round_up(n, decimals=-1):
        multiplier = 10 ** decimals
        return math.ceil(n * multiplier) / multiplier

    extreme1 = round_up(max([abs(df.npv_sum.min()), abs(df.npv_sum.max())]) / 10 ** 6, -3)
    extreme2 = round_up(max([abs(df.cashflow.min()), abs(df.cashflow.max())]) / 10 ** 6, -3)
    extreme3 = round_up(max([abs(df.capex.min()), abs(df.capex.max())]) / 10 ** 6, -3)
    extreme4 = round_up(max([abs(df.opex.min()), abs(df.opex.max())]) / 10 ** 6, -3)
    extreme5 = round_up(max([abs(df.revenue.min()), abs(df.revenue.max())]) / 10 ** 6, -3)
    extreme = max([extreme1, extreme2, extreme3, extreme4, extreme5])

    if not cash_flow_lims:
        cash_flow_lims = [-1.1*extreme, 1.1*extreme]

    if not npv_lims:
        npv_lims = [-1.1*extreme, 1.1*extreme]

    # assert that dataframe adheres to prescribed standards
    test_dataframe(df)

    # preset fontsize and legend fontsize
    fontsize = 20
    fontsize_legend = 15

    # initialise figure
    fig, ax1 = plt.subplots(1, 1, sharex=True, figsize=(16, 8))
    plt.rcParams['font.size'] = fontsize

    offset = 0.25
    width = 0.25

    plt.axis('off')

    ax1 = fig.add_subplot(1, 1, 1)

    # ----

    ax1.bar([x - 1 * offset for x in df['years']], height=list(df['capex'] / 10 ** 6), color='red', width=width,
            label='CAPEX')
    ax1.bar([x + 0 * offset for x in df['years']], height=list(df['opex'] / 10 ** 6), color='blue', width=width,
            label='OPEX')
    ax1.bar([x + 1 * offset for x in df['years']], height=list(df['revenue'] / 10 ** 6), color='green', width=width,
            label='Revenue')

    ax1.legend(loc='lower left', ncol=3, fontsize=fontsize_legend, bbox_to_anchor=(x1, y1), frameon=False)

    ax1.set_title(title, fontsize=fontsize)
    ax1.set_xlabel(r'Years', fontsize=fontsize, labelpad=20)
    ax1.set_ylabel(r'Cash flows ($10^6$ Euro)', fontsize=fontsize)
    ax1.set_xticks(np.arange(0, max(df['years']) + 1, 1))
    ax1.set_xticklabels(['{:.0f}'.format(x) for x in np.arange(0, max(df['years']) + 1, 1)], rotation=90,
                        fontsize=fontsize)

    ax1.grid(which='major', axis='both')
    ax1.set_xlim([df.years.min() - 1, df.years.max() + 1])
    ax1.set_ylim(cash_flow_lims)
    # ----

    ax2 = ax1.twinx()  # instantiate a second axes that shares the same x-axis

    ax2.plot(list(df['years']), list(df['cashflow_sum'] / 10 ** 6), color='red', marker='o', label='Cumulative cashflows (Capex, Opex, Revenues)')

    ax2.legend(loc='lower right', fontsize=fontsize_legend, bbox_to_anchor=(x2 + 1, y2), frameon=False)

    ax2.set_ylabel('NPV ($10^6$ Euro)', fontsize=fontsize)  # we already handled the x-label with ax1

    ax2.set_ylim(npv_lims)  # NB: you want to take care that the y=0 of ax1 and ax2 align to avoid confusion

    # ----
    
    # add a marker in the final year representing the cumulated NPV value
    final_year = df['years'].iloc[-1]
    final_npv = df['npv_sum'].iloc[-1]
    ax2.scatter(final_year, final_npv / 10 ** 6, color='purple', s=500, marker='x', label='Cumulative NPV in final year')
    ax2.legend(loc='lower right', fontsize=fontsize_legend, bbox_to_anchor=(x2 + 1, y2), frameon=False)
    
    fig.tight_layout()

    fig.tight_layout = True


def load_input_from_xls(filename=r'H2 Model - Input sheet.xlsm'):
    # load workbook
    wb = load_workbook(filename)

    # load worksheet
    ws = wb["Input Tab"]

    mapping = {}

    for entry, data_boundary in ws.tables.items():
        # parse the data within the ref boundary
        data = ws[data_boundary]

        # extract the data (the inner list comprehension gets the values for each cell in the table)
        content = [[cell.value for cell in ent] for ent in data]

        # find header
        header = content[0]

        # find the rest ... the contents, excluding the header
        rest = content[1:]

        # create dataframe with the column names
        # and pair table name with dataframe
        df = pd.DataFrame(rest, columns=header)
        mapping[entry] = df

    return mapping


def extract_scenario(mapping, scenario):
    return pd.DataFrame(mapping[scenario])


def get_object_data(Inputs,
                    subsystem='Wind energy source & Transport',
                    element='Offshore wind park',
                    component='Foundations',
                    capex_categories=['Development and Project Management', 'Procurement',
                                      'Installation and Commissioning'],
                    opex_categories=['Yearly Variable Costs Rate', 'Insurance Rate'],
                    debug=False):
    """
    Assuming columns Sub-system, Element and Component allways exist

    Method returns cashflow dataframe
    """

    object_data = {
        'sub_system': subsystem,
        'element': element,
        'component': component,
        'escalation_base_year': [],
        'escalation_rate': [],
        'capex_per_unit': [],
        'capex_per_unit_units': [],
        'unit': [],
        'unit_units': [],
        'construction_duration': [],
        'share_of_investments': [],
        'economic_lifetime': [],
        'depreciation_flag': [],
        'depreciation_rate': [],
        'yearly_variable_costs_flag': [],
        'yearly_variable_costs_rate': [],
        'insurance_flag': [],
        'insurance_rate': [],
        'decommissioning_rate': [],
        'residual_value': []}


    # Escalation base year
    try:
        object_data['escalation_base_year'] = int(Inputs[
            (Inputs['Category'] == 'System input') &
            (Inputs['Description'].str.contains('Escalation base year'))].Number.item())

    except:
        print('issue detected')

    # Escalation rate
    try:
        object_data['escalation_rate'] = Inputs[
            (Inputs['Category'] == 'System input') &
            (Inputs['Description'].str.contains('Escalation rate'))].Number.item()

    except:
        print('issue detected')

    # capex_per_unit
    try:
        object_data['capex_per_unit'] = Inputs[
            (Inputs['Sub-system'] == subsystem) &
            (Inputs['Element'] == element) &
            (Inputs['Component'] == component) &
            (Inputs['Description'].str.contains(
                '|'.join(capex_categories)))].Number.sum()

        object_data['capex_per_unit_units'] = Inputs[
            (Inputs['Sub-system'] == subsystem) &
            (Inputs['Element'] == element) &
            (Inputs['Component'] == component) &
            (Inputs['Description'].str.contains(
                '|'.join(capex_categories)))].Unit.unique().item()

    except:
        print('issue detected - wow')

    # Number of units
    try:
        object_data['unit'] = Inputs[
            (Inputs['Sub-system'] == subsystem) &
            (Inputs['Element'] == element) &
            (Inputs['Component'] == component) &
            (Inputs['Description'] == 'Number of units')].Number.item()

    except:
        print('issue detected')

    # Unit units
    try:
        object_data['unit_units'] = Inputs[
            (Inputs['Sub-system'] == subsystem) &
            (Inputs['Element'] == element) &
            (Inputs['Component'] == component) &
            (Inputs['Description'] == 'Number of units')].Unit.item()

    except:
        print('issue detected')

    # Construction duration (must be an integer)
    try:
        object_data['construction_duration'] = int(Inputs[
                                                       (Inputs['Sub-system'] == subsystem) &
                                                       (Inputs['Element'] == element) &
                                                       (Inputs['Component'] == component) &
                                                       (Inputs[
                                                            'Description'] == 'Construction duration')].Number.item())

    except:
        print('issue detected')

    # Share of Investments
    try:
        # isolate the rows that contain 'Share of Investments in Year' and remove the string to only get the year numbers
        years = list(Inputs[
            (Inputs['Sub-system'] == subsystem) &
            (Inputs['Element'] == element) &
            (Inputs['Component'] == component) &
            (Inputs['Description'].str.contains('Share of Investments in Year'))].Description.str.replace(
            'Share of Investments in Year ', ''))

        # transform the year numbers from string to in and sort them to be certain of the order
        years = [int(x) for x in years]
        years.sort()

        # now extract the allocations, since years are sorted we know for sure now that the allocations are in the right order
        Construction_allocation = []
        for year in years:
            Construction_allocation.append(Inputs[
                                               (Inputs['Sub-system'] == subsystem) &
                                               (Inputs['Element'] == element) &
                                               (Inputs['Component'] == component) &
                                               (Inputs['Description'].str.contains(
                                                   'Share of Investments in Year ' + str(year)))
                                               ].Number.item())

        object_data['share_of_investments'] = Construction_allocation[:object_data['construction_duration']]

    except:
        print('issue detected')

    # Economic Lifetime (must be an integer)
    try:
        object_data['economic_lifetime'] = int(Inputs[
                                                   (Inputs['Sub-system'] == subsystem) &
                                                   (Inputs['Element'] == element) &
                                                   (Inputs['Component'] == component) &
                                                   (Inputs['Description'] == 'Economic Lifetime')].Number.item())

    except:
        print('issue detected')

    # Depreciation Flag
    try:
        object_data['depreciation_flag'] = int(Inputs[
            (Inputs['Sub-system'] == subsystem) &
            (Inputs['Element'] == element) &
            (Inputs['Component'] == component) &
            (Inputs['Description'] == 'Depreciation Flag')].Number.item())

    except:
        print('issue detected')

    # Depreciation rate
    try:
        object_data['depreciation_rate'] = Inputs[
            (Inputs['Sub-system'] == subsystem) &
            (Inputs['Element'] == element) &
            (Inputs['Component'] == component) &
            (Inputs['Description'] == 'Depreciation Rate')].Number.item()

    except:
        print('issue detected')

    # Yearly Variable Costs Flag
    try:
        object_data['yearly_variable_costs_flag'] = int(Inputs[
            (Inputs['Sub-system'] == subsystem) &
            (Inputs['Element'] == element) &
            (Inputs['Component'] == component) &
            (Inputs['Description'] == 'Yearly Variable Costs Flag')].Number.item())

    except:
        print('issue detected')

    # Yearly Variable Costs Rate
    try:
        object_data['yearly_variable_costs_rate'] = Inputs[
            (Inputs['Sub-system'] == subsystem) &
            (Inputs['Element'] == element) &
            (Inputs['Component'] == component) &
            (Inputs['Description'] == 'Yearly Variable Costs Rate')].Number.item()

    except:
        print('?issue detected')

    # Insurance Flag
    try:
        object_data['insurance_flag'] = int(Inputs[
            (Inputs['Sub-system'] == subsystem) &
            (Inputs['Element'] == element) &
            (Inputs['Component'] == component) &
            (Inputs['Description'] == 'Insurance Flag')].Number.item())

    except:
        print('issue detected')

    # Insurance Rate
    try:
        object_data['insurance_rate'] = Inputs[
            (Inputs['Sub-system'] == subsystem) &
            (Inputs['Element'] == element) &
            (Inputs['Component'] == component) &
            (Inputs['Description'] == 'Insurance Rate')].Number.item()

    except:
        print('issue detected')

    # Decommissioning
    try:
        object_data['decommissioning_rate'] = Inputs[
            (Inputs['Sub-system'] == subsystem) &
            (Inputs['Element'] == element) &
            (Inputs['Component'] == component) &
            (Inputs['Description'].str.contains('decommissioning'))].Number.item()

    except:
        print('issue detected')

    # Residual Value
    try:
        object_data['residual_value'] = Inputs[
            (Inputs['Sub-system'] == subsystem) &
            (Inputs['Element'] == element) &
            (Inputs['Component'] == component) &
            (Inputs['Description'].str.contains('residual value'))].Number.item()

    except:
        print('issue detected')

    return object_data
